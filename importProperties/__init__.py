import azure.functions as func
import logging
import pyodbc
import openpyxl
import io

username = 'dylanwalls'
password = '950117Dy!'
server = 'scorecard-server.database.windows.net'
database = 'dashboard-new-server'
driver = '{ODBC Driver 17 for SQL Server}'

def import_property_data(excel_data):
    try:
        # Connect to your Azure SQL Database
        conn = pyodbc.connect(
            f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
        )
        cursor = conn.cursor()

        # Load the uploaded Excel file from BytesIO
        workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
        worksheet = workbook.active

        logging.info('Starting Excel file processing...')
        # for row in worksheet.iter_rows(min_row=2, values_only=True):
        #     logging.info(row)

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            homeowner = row[0]  # Modify this to match the column containing homeowner
            street = row[1]  # Modify this to match the column containing street
            property_ref = row[2]  # Modify this to match the column containing property_ref
            no_units = row[3]  # Modify this to match the column containing no_units
            rent = row[4]  # Modify this to match the column containing rent
            suburb = row[5]  # Modify this to match the column containing suburb
            comments = row[6]  # Modify this to match the column containing manual_comments

            # Check if the property already exists in the database
            cursor.execute('SELECT property_ref FROM Properties WHERE property_ref = ?', (property_ref,))
            existing_property = cursor.fetchone()

            # If the property does not exist, insert it into the database
            if existing_property is None:
                logging.info(f'No property exists: {property_ref}')
                cursor.execute('''
                    INSERT INTO Properties (property_ref, homeowner, street, suburb, no_units, rent)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (property_ref, homeowner, street, suburb, no_units, rent))
                logging.info(f'Inserted property: {property_ref}')

            # Generate rental unit data based on the 'no_units' value
            for unit_num in range(1, no_units + 1):
                unit_letter = chr(96 + unit_num)  # Convert unit number to corresponding letter
                unit_ref = f'{property_ref}{unit_letter}'
                rent = rent  # You can modify this if needed
                deposit_due = rent
                # logging.info(f'Processing unit: {unit_ref}')

                # Check if the rental unit already exists in the database
                cursor.execute('SELECT unit_ref FROM RentalUnits WHERE unit_ref = ?', (unit_ref,))
                existing_rental_unit = cursor.fetchone()

                # If the rental unit does not exist, insert it into the database
                if existing_rental_unit is None:
                    logging.info(f'No existing rental unit: {unit_ref}')
                    cursor.execute('SELECT property_id FROM Properties WHERE property_ref = ?', (property_ref,))
                    property_id = cursor.fetchone()[0]
                    logging.info(f'property_id is: {property_id}')
                    cursor.execute('''
                        INSERT INTO RentalUnits (property_id, unit_ref, rent, deposit_due)
                        VALUES (?, ?, ?, ?)
                    ''', (property_id, unit_ref, rent, deposit_due))
                    logging.info(f'Inserted rental unit: {unit_ref}')

        # Commit the changes and close the connection
        conn.commit()
        conn.close()

        logging.info('Excel file processing completed successfully.')
        return "Property data imported successfully."

    except Exception as e:
        logging.error(f'An error occurred: {str(e)}')
        return f"An error occurred: {str(e)}"

def main(req: func.HttpRequest) -> func.HttpResponse:
    try:
        logging.basicConfig(level=logging.INFO)

        # Check if 'excelFile' is in the files dictionary
        if 'excelFile' in req.files:
            # Get the uploaded file
            file = req.files['excelFile']

            # Read the uploaded file content as bytes
            file_content = file.read()

            result = import_property_data(file_content)

            return func.HttpResponse(result, status_code=200)
        else:
            # If 'excelFile' is not in the files dictionary, return a bad request response
            logging.error("Missing 'excelFile' in request.")
            return func.HttpResponse("Missing 'excelFile' in request.", status_code=400)

    except Exception as e:
        logging.error(f'An error occurred: {str(e)}')
        return func.HttpResponse(f"An error occurred: {str(e)}", status_code=500)
