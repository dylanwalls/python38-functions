import azure.functions as func
import logging
import pyodbc
import openpyxl
import io

username = 'dylanwalls'
password = '950117Dy!'
server = 'scorecard-server.database.windows.net'
database = 'dashboard-new-server'
driver = '{ODBC Driver 17 for SQL Server}'

def import_meter_data(excel_data):
    try:
        # Connect to your Azure SQL Database
        conn = pyodbc.connect(
            f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}"
        )
        cursor = conn.cursor()

        # Load the uploaded Excel file from BytesIO
        workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
        worksheet = workbook.active

        logging.info('Starting Excel file processing...')
        # for row in worksheet.iter_rows(min_row=2, values_only=True):
        #     logging.info(row)

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            Building = row[0]  # Modify this to match the column containing homeowner
            Unit = row[1]  # Modify this to match the column containing street
            # Desc = row[2]  # Modify this to match the column containing property_ref
            Meter = str(row[3])  # Modify this to match the column containing no_units
            # ACI = row[4]  # Modify this to match the column containing rent
            Type = row[5]  # Modify this to match the column containing suburb
            SGC = row[6]  # Modify this to match the column containing manual_comments
            # Install = row[7]
            Blocked = row[8]
            UnitID = row[9]

            # Check if the property already exists in the database
            cursor.execute('SELECT * FROM citiqMeters WHERE Meter = ?', (Meter,))
            existing_meter = cursor.fetchone()

            # If the property does not exist, insert it into the database
            if existing_meter is None:
                logging.info(f'No meter exists: {Meter}')
                cursor.execute('''
                    INSERT INTO citiqMeters (Building, Unit, Meter, Type, SGC, Blocked, [Unit ID])
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (Building, Unit, Meter, Type, SGC, Blocked, UnitID))
                logging.info(f'Inserted meter: {Meter}')

        # Commit the changes and close the connection
        conn.commit()
        conn.close()

        logging.info('Excel file processing completed successfully.')
        return "Meter data imported successfully."

    except Exception as e:
        logging.error(f'An error occurred: {str(e)}')
        return f"An error occurred: {str(e)}"

def main(req: func.HttpRequest) -> func.HttpResponse:
    try:
        logging.basicConfig(level=logging.INFO)

        # Check if 'excelFile' is in the files dictionary
        if 'excelFile' in req.files:
            # Get the uploaded file
            file = req.files['excelFile']

            # Read the uploaded file content as bytes
            file_content = file.read()

            result = import_meter_data(file_content)

            return func.HttpResponse(result, status_code=200)
        else:
            # If 'excelFile' is not in the files dictionary, return a bad request response
            logging.error("Missing 'excelFile' in request.")
            return func.HttpResponse("Missing 'excelFile' in request.", status_code=400)

    except Exception as e:
        logging.error(f'An error occurred: {str(e)}')
        return func.HttpResponse(f"An error occurred: {str(e)}", status_code=500)
